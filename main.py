import openpyxl
inventory_file=openpyxl.load_workbook("inventory.xlsx")
product_list=inventory_file["Sheet1"]
products_per_supplier = {}
product_inventory_less_than_10 = []
inventory_value_for_each_company = {}
print(product_list.max_row)

#calculation for no f products per supplier
for product_row in range(2,product_list.max_row + 1):
    supplier_name = product_list.cell(product_row,4).value
   
    if supplier_name in products_per_supplier:
       
       products_per_supplier[supplier_name]=products_per_supplier[supplier_name]+ 1
    else:
        print("Adding new supp")
        products_per_supplier[supplier_name] = 1
    if product_list.cell(product_row,2).value<10:
       product_inventory_less_than_10.append(product_row)
    company_name = product_list.cell(product_row,4).value
    inventory_value =  product_list.cell(product_row,2).value * product_list.cell(product_row,3).value
   
    if company_name in inventory_value_for_each_company:
       
       inventory_value_for_each_company[company_name]=inventory_value_for_each_company[company_name] + inventory_value
    else:
        print("Adding new inventory list")
        inventory_value_for_each_company[company_name] = inventory_value

    product_list.cell(product_row,5).value=product_list.cell(product_row,2).value * product_list.cell(product_row,3).value
   
    inventory_file.save("new_inventory_file.xlsx")
        
print(product_inventory_less_than_10)
print(products_per_supplier)
print(inventory_value_for_each_company)
   

#list less than 10

   
#total inventory value

    

